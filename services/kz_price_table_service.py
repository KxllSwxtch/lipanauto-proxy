"""
KZ Price Table Service
Parses kz-table.xlsx file for Kazakhstan customs price lookups
"""
import os
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class KZPriceTableService:
    """
    Service to parse and query the Kazakhstan price table

    The kz-table.xlsx file contains car prices used for Kazakhstan customs calculations
    with columns: Марка, Модель, Объём, Год выпуска, Стоимость доллар США
    """

    def __init__(self, file_path: str = None):
        """
        Initialize the service and load the price table

        Args:
            file_path: Path to kz-table.xlsx file
        """
        if file_path is None:
            # Default path (one level up from lipanauto-proxy)
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            file_path = os.path.join(base_dir, "kz-table.xlsx")

        self.file_path = file_path
        self.price_data: List[Dict] = []
        self.is_loaded = False

        # Load the table on initialization
        self._load_price_table()

    def _load_price_table(self):
        """Load and parse the Excel file into memory"""
        try:
            if not os.path.exists(self.file_path):
                print(f"⚠️  KZ price table not found at: {self.file_path}")
                self.is_loaded = False
                return

            print(f"📊 Loading KZ price table from: {self.file_path}")

            # Load workbook in read-only mode for efficiency
            wb = load_workbook(filename=self.file_path, read_only=True, data_only=True)

            # Get the first sheet (or you can specify the sheet name if known)
            ws = wb.active

            # Parse headers (first row)
            headers = []
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(header_row)]

            print(f"📋 Found columns: {headers}")

            # Detect column indices (flexible column detection)
            col_indices = self._detect_columns(headers)

            if not all(col_indices.values()):
                print(f"⚠️  Could not detect all required columns. Found: {col_indices}")
                wb.close()
                self.is_loaded = False
                return

            # Parse data rows (skip header)
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    # Extract values based on detected columns
                    manufacturer = str(row[col_indices['manufacturer']]).strip() if row[col_indices['manufacturer']] else None
                    model = str(row[col_indices['model']]).strip() if row[col_indices['model']] else None
                    volume = row[col_indices['volume']]
                    year = row[col_indices['year']]
                    price_usd = row[col_indices['price_usd']]

                    # Skip rows with missing critical data
                    if not all([manufacturer, model, volume, year, price_usd]):
                        continue

                    # Convert to appropriate types
                    # Round volume to 1 decimal place to match table format (e.g., 2.998 → 3.0)
                    volume = round(float(volume), 1) if volume else None
                    year = int(year) if year else None
                    price_usd = float(price_usd) if price_usd else None

                    # Add to price data
                    self.price_data.append({
                        'manufacturer': manufacturer.lower(),  # Normalize for matching
                        'model': model.lower(),
                        'volume': volume,
                        'year': year,
                        'price_usd': price_usd,
                        'manufacturer_original': manufacturer,
                        'model_original': model
                    })

                    row_count += 1

                except Exception as e:
                    # Skip problematic rows
                    continue

            wb.close()

            print(f"✅ Loaded {row_count} entries from KZ price table")
            self.is_loaded = True

        except Exception as e:
            print(f"❌ Error loading KZ price table: {e}")
            self.is_loaded = False

    def _detect_columns(self, headers: List[str]) -> Dict[str, Optional[int]]:
        """
        Detect column indices based on header names

        Looks for keywords in headers (case-insensitive, partial matches)

        Returns:
            Dict with keys: manufacturer, model, volume, year, price_usd
        """
        col_indices = {
            'manufacturer': None,
            'model': None,
            'volume': None,
            'year': None,
            'price_usd': None
        }

        for i, header in enumerate(headers):
            header_lower = header.lower()

            # Manufacturer: марка, brand, manufacturer, make
            if 'марка' in header_lower or 'brand' in header_lower or 'manufacturer' in header_lower or 'make' in header_lower:
                col_indices['manufacturer'] = i

            # Model: модель, model
            elif 'модель' in header_lower or 'model' in header_lower:
                col_indices['model'] = i

            # Volume: объём, объем, volume, displacement, cc
            elif 'объём' in header_lower or 'объем' in header_lower or 'volume' in header_lower or 'displacement' in header_lower or 'cc' in header_lower:
                col_indices['volume'] = i

            # Year: год, year
            elif 'год' in header_lower or 'year' in header_lower:
                col_indices['year'] = i

            # Price USD: стоимость, price, usd, доллар
            elif 'стоимость' in header_lower or ('price' in header_lower and 'usd' in header_lower) or 'доллар' in header_lower:
                col_indices['price_usd'] = i

        return col_indices

    def lookup_price(
        self,
        manufacturer: str,
        model: str,
        volume: float,
        year: int,
        volume_tolerance: float = 0.2
    ) -> Optional[float]:
        """
        Lookup car price in USD from the KZ price table

        Args:
            manufacturer: Car manufacturer (e.g., "Hyundai", "Kia")
            model: Car model (e.g., "Sonata", "K5")
            volume: Engine volume in liters (e.g., 2.0)
            year: Manufacturing year (e.g., 2020)
            volume_tolerance: Tolerance for volume matching (default 0.2L)

        Returns:
            Price in USD or None if not found
        """
        if not self.is_loaded:
            print("⚠️  KZ price table not loaded")
            return None

        # Normalize inputs
        manufacturer_lower = manufacturer.lower().strip()
        model_lower = model.lower().strip()

        # Round volume to 1 decimal place to match table format
        original_volume = volume
        volume = round(volume, 1)

        if original_volume != volume:
            print(f"📊 Rounded engine volume: {original_volume}L → {volume}L for lookup")

        # Search for matching entries
        best_match = None
        best_match_score = 0

        for entry in self.price_data:
            # Exact manufacturer and model match
            if entry['manufacturer'] != manufacturer_lower:
                continue

            if entry['model'] != model_lower:
                continue

            # Volume match with tolerance
            if entry['volume']:
                volume_diff = abs(entry['volume'] - volume)
                if volume_diff > volume_tolerance:
                    continue

            # Year match (exact)
            if entry['year'] != year:
                continue

            # Found an exact match
            return entry['price_usd']

        # If no exact match, try relaxed matching
        return self._fuzzy_lookup(manufacturer_lower, model_lower, volume, year, volume_tolerance)

    def _fuzzy_lookup(
        self,
        manufacturer: str,
        model: str,
        volume: float,
        year: int,
        volume_tolerance: float = 0.2
    ) -> Optional[float]:
        """
        Fuzzy matching with relaxed constraints

        Tries:
        1. Same manufacturer, model, volume with ±1 year
        2. Same manufacturer, model with closest year
        3. Fallback to None
        """
        # Round volume to 1 decimal place for consistency
        volume = round(volume, 1)

        # Try ±1 year
        for year_offset in [-1, 1]:
            adjusted_year = year + year_offset
            for entry in self.price_data:
                if (entry['manufacturer'] == manufacturer and
                    entry['model'] == model and
                    entry['year'] == adjusted_year):

                    if entry['volume']:
                        volume_diff = abs(entry['volume'] - volume)
                        if volume_diff <= volume_tolerance:
                            print(f"📌 Found fuzzy match: {year_offset:+d} year offset")
                            return entry['price_usd']

        # Try closest year match
        closest_entry = None
        min_year_diff = float('inf')

        for entry in self.price_data:
            if (entry['manufacturer'] == manufacturer and
                entry['model'] == model):

                if entry['volume']:
                    volume_diff = abs(entry['volume'] - volume)
                    if volume_diff > volume_tolerance:
                        continue

                year_diff = abs(entry['year'] - year)
                if year_diff < min_year_diff:
                    min_year_diff = year_diff
                    closest_entry = entry

        if closest_entry:
            print(f"📌 Found closest match: {min_year_diff} years difference")
            return closest_entry['price_usd']

        print(f"❌ No match found for {manufacturer} {model} {volume}L {year}")
        return None

    def get_available_manufacturers(self) -> List[str]:
        """Get list of all manufacturers in the table"""
        if not self.is_loaded:
            return []

        manufacturers = set(entry['manufacturer_original'] for entry in self.price_data)
        return sorted(list(manufacturers))

    def get_models_by_manufacturer(self, manufacturer: str) -> List[str]:
        """Get list of models for a specific manufacturer"""
        if not self.is_loaded:
            return []

        manufacturer_lower = manufacturer.lower().strip()
        models = set(
            entry['model_original']
            for entry in self.price_data
            if entry['manufacturer'] == manufacturer_lower
        )
        return sorted(list(models))

    def reload_table(self):
        """Reload the price table from disk"""
        self.price_data = []
        self._load_price_table()


# Global singleton instance
kz_price_table_service = KZPriceTableService()
